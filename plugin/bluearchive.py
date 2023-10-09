import discord
from discord.ext import commands
from discord.ext.commands import Bot
from discord import app_commands
from discord.ext.commands import check, Context
from typing import List
from discord import ButtonStyle, ActionRow
from discord.components import Button

import traceback
import sys
import datetime
import math
import copy
import asyncio
from itertools import groupby

import utils.logger as logger
from utils.database.connection import select, raw, execute_script
from staticrunner import StaticRunner as sr
import utils.api.api as api
from utils.cutils import get_value_from_nested_dict, add_to_nested_dict, default_if_empty, remove_nested_item, generate_excel as gen_excel, get_date, check_similar
from utils.discord.utils import context, response

import utils.paginator as page
import pandas as pd
# import img2pdf
import openpyxl

class BlueArchive(commands.Cog):
    """Blue Archive utilities"""

    __slots__ = ('bot', 'students_option', 'schools_option')

    def __init__(self, bot):
        self.bot = bot
        self.logger = logger.Logger("BlueArchive")

        self.logger.debug("==== initiate options ====")
        self.schools_option = self.fetch_school()
        self.students_option = self.fetch_students()

        whitelist = self.fetch_whitelist()
        clubs = self.fetch_club()
        self.logger.debug("==== initiate options ====")

    async def __local_check(self, ctx):
        """A local check which applies to all commands in this cog."""
        if not ctx.guild:
            raise commands.NoPrivateMessage
        return True

    async def __error(self, ctx, error):
        """A local error handler for all errors arising from commands in this cog."""
        if isinstance(error, commands.NoPrivateMessage):
            try:
                return await response(ctx,'這個指令無法在私訊使用哦~')
            except discord.HTTPException:
                pass

        print('Ignoring exception in command {}:'.format(ctx.command), file=sys.stderr)
        traceback.print_exception(type(error), error, error.__traceback__, file=sys.stderr)

    @commands.Cog.listener()
    async def on_command_error(self, ctx: commands.Context, error: Exception): 
        try:
            await ctx.message.delete(delay=10)
        except:
            traceback.print_exception(type(error), error, error.__traceback__, file=sys.stderr)
    
    # region !CRITICAL CHECKING! Check permited roles, in default only server owner, admin and bot owner
    async def permitted_to_execute(self, ctx: commands.Context, club: int = 0):
        return set(default_if_empty(get_value_from_nested_dict(sr.ba_club_manage_whitelist_roles, [ctx.guild.id, club]), [])) & set([role.id for role in ctx.author.roles]) or self.owner_only(ctx) or ctx.author.id in get_value_from_nested_dict(sr.ba_club_manage_whitelist_members, [ctx.guild.id, club])
    
    def owner_only(self, ctx: commands.Context):
        return ctx.author in [self.bot.application.owner, ctx.guild.owner]
    
    def admin_or_owner():
        async def predicate(ctx: Context):
            return ctx.author in [ctx.guild.owner] or ctx.message.author.guild_permissions.administrator
        return check(predicate)
    # endregion
    
    # region Fetch data from database
    def fetch_school(self, force = False):
        if sr.ba_school_names is None or force:
            r = select('BA_SCHOOL', ('NAME',))[2]
            sr.ba_school_names = [d['NAME'] for d in r if 'NAME' in d]
        return sr.ba_school_names
    
    def fetch_students(self, force = False) -> list[tuple]:
        if sr.ba_characters is None or force:
            r = select('BA_CHARACTER', ('ID', 'C_NAME', 'C_SCHOOL'))[2]
            sr.ba_characters = [(d['ID'], d['C_NAME'], d['C_SCHOOL']) for d in r]
            sr.ba_character_names = [d['C_NAME'] for d in r]
        return sr.ba_characters
    
    def fetch_whitelist(self, force = False) -> tuple:
        whitelist_members = sr.ba_club_manage_whitelist_members
        whitelist_roles = sr.ba_club_manage_whitelist_roles
        if len(whitelist_members) == 0 or len(whitelist_roles) == 0 or force:
            sr.ba_club_manage_whitelist_members = {}
            sr.ba_club_manage_whitelist_roles = {}
            result = select('BA_CLUB_MANAGE_WHITELIST', ('ID', 'GUILD_ID', 'SERVER_ID', 'TYPE'))[2]
            for r in result:
                if r.get('TYPE') == 1:
                    add_to_nested_dict(sr.ba_club_manage_whitelist_members, [r.get('SERVER_ID'), r.get('GUILD_ID')], r.get('ID'), as_list = True)
                elif r.get('TYPE') == 2:
                    add_to_nested_dict(sr.ba_club_manage_whitelist_roles, [r.get('SERVER_ID'), r.get('GUILD_ID')], r.get('ID'), as_list = True)
                else:
                    self.logger.warning("Invalid type code")
            self.logger.debug(sr.ba_club_manage_whitelist_members)
            self.logger.debug(sr.ba_club_manage_whitelist_roles)
        return (sr.ba_club_manage_whitelist_members, sr.ba_club_manage_whitelist_roles)
    
    def fetch_club(self, force = False) -> dict:
        ba_clubs = sr.ba_clubs
        if len(ba_clubs) == 0 or force:
            result = select('BA_GUILD', ('SERVER_ID', 'GUILD_ID', 'GUILD_NAME', 'GUILD_DESCRIPTION', 'DEFAULT_ROLE', 'CONFIGURED_BY'))[2]
            for r in result:
                add_to_nested_dict(sr.ba_clubs, [r.get('SERVER_ID'), r.get('GUILD_ID')], 
                                   {'name':r.get('GUILD_NAME'), 'desc':r.get('GUILD_DESCRIPTION'), 'd_role':r.get('DEFAULT_ROLE'), 'initial':r.get('CONFIGURED_BY')})
            self.logger.debug(sr.ba_clubs)
        return sr.ba_clubs

    def get_my_student(self, user) -> list[str]:
        r, result = raw(f"SELECT C_NAME FROM BA_MEMBER_STUDENTS WHERE MEMBER_ID = '{user}'")[1:]
        if not r == 0:
            return [std.get('C_NAME') for std in result]
        return ["There has no student in your list"]
    
    def fetch_member_by_server(self, server_id) -> list[tuple]:
        r, result = raw(f"SELECT MEMBER_ID, MEMBER_IGN FROM BA_GUILD_MEMBER WHERE SERVER_ID = '{server_id}'")[1:]
        if not r == 0:
            return [(std.get('MEMBER_IGN'), std.get('MEMBER_ID')) for std in result]
        return None

    # endregion

    # region Options/AutoComplete
    async def school_options(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        return [
            app_commands.Choice(name=choice, value=choice)
            for choice in sr.ba_school_names if current.lower() in choice.lower()
        ]
    
    async def student_options(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        filter = [app_commands.Choice(name=f"All student", value="*")]
        if current is not None and not current == '':
            filter = [
                app_commands.Choice(name=f"{choice[1]}, {choice[2]}", value=str(choice[1]))
                for choice in sr.ba_characters if str(choice[1]).lower().startswith(current.lower())
            ]
        return filter
    
    async def rarity(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        choices = sr.ba_student_rarity[4:]
        return [
            app_commands.Choice(name=choice, value=choice)
            for choice in choices if current.lower() in choice.lower()
        ]
    
    async def rarity_full(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        choices = sr.ba_student_rarity
        return [
            app_commands.Choice(name=choice, value=choice)
            for choice in choices if current.lower() in choice.lower()
        ]
    
    async def reload_options(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        choices = ['all', 'student', 'school']
        return [
            app_commands.Choice(name=choice, value=choice)
            for choice in choices if current.lower() in choice.lower()
        ]
    
    async def skill_level_options(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        option: str = list(filter(None, [d.get('name') if d.get('focused') else None for d in ctx.data.get('options')[0].get('options')]))[0]
        choices = [str(i + 1) for i in range(9)]
        if 'ex' in option:
            choices = choices[:4]
        choices.append('M')
        choices.reverse()
        return [
            app_commands.Choice(name=choice, value=choice)
            for choice in choices if current.lower() in choice.lower()
        ]
    
    async def manage_option(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        choices = ['Add', 'Update','Remove', 'View']
        return [
            app_commands.Choice(name=choice, value=choice)
            for choice in choices if current.lower() in choice.lower()
        ]
    
    async def whitelist_group_options(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        choices = ['member', 'role']
        return [
            app_commands.Choice(name=choice, value=choice)
            for choice in choices if current.lower() in choice.lower()
        ]
    
    async def clubs_options(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[int]]:
        choices = [(k, v.get('name')) for k, v in sr.ba_clubs.get(ctx.guild.id).items()]
        return [
            app_commands.Choice(name=choice[1], value=choice[0])
            for choice in choices if current in choice[1]
        ]
    
    async def member_type_options(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        choices = ['President', 'Vice', 'Manager', 'Member']
        return [
            app_commands.Choice(name=choice, value=choice)
            for choice in choices if current.lower() in choice.lower()
        ]
    
    async def club_member_option(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[int]]:
        # option: str = list(filter(None, [d.get('name') if d.get('focused') else None for d in ctx.data.get('options')[0].get('options')]))[0]
        
        member = self.fetch_member_by_server(ctx.guild.id)
        filter = [app_commands.Choice(name=f"There has no member registred in club or no club exist in this server", value=0)]
        if member:
            filter = [app_commands.Choice(name=f"Start typing name of member for suggestion", value=0)]
            if len(member) <= 10 or current is not None and not current == '':
                filter = [
                    app_commands.Choice(name=ign, value=int(member_id))
                    for ign, member_id in member if current.lower() in ign.lower()
                ]
        return filter 
    
    async def mystudent_options(self,
        ctx: discord.Interaction|commands.Context,
        current: str,
    ) -> List[app_commands.Choice[str]]:
        students = self.get_my_student(ctx.user.id)
        filter = [app_commands.Choice(name=f"Start typing name of student for suggestion", value="*")]
        if len(students) <= 10 or current is not None and not current == '':
            filter = [
                app_commands.Choice(name=choice, value=choice)
                for choice in students if choice.lower().startswith(current.lower())
            ]
        return filter 
    # endregion
    
    # region this part is for owner to reload the configuration or database update
    @app_commands.command(name='reload', description='owner only access to reload the configuration')
    @app_commands.autocomplete(option=reload_options)
    @app_commands.describe(option="The item to be reloaded")
    @app_commands.describe(reload_database="Default true, Whether to reload the database or not")
    @app_commands.describe(generate_excel="Default true, Whether to generate excel list")
    @app_commands.rename(reload_database="reload-database")
    @app_commands.rename(generate_excel="generate-excel")
    @commands.is_owner()
    async def reload_configuration(self, ctx: discord.Interaction|commands.Context, option: str = 'all', reload_database: bool = True, generate_excel: bool = True):
        """Show the student details"""
        self.logger.debug("===== reload-configuration - start =====")
        self.logger.debug(f"option => {option}")

        try:
            ctx = await context(ctx, ephemeral=True)

            msg = await response(ctx,embed=discord.Embed(color=discord.Color.from_str('#02D3FB'), description="start reload configuration"), ephemeral=True)

            if option == 'school':
                api.BlueArhieveAPI.fetch_ba_school_list(reload_database)
                self.fetch_school(True)
            elif option == 'student':
                api.BlueArhieveAPI.fetch_ba_character_list(reload_database)
                self.fetch_students(True)
            else:
                api.BlueArhieveAPI.fetch_ba_school_list(reload_database)
                self.fetch_school(True)
                api.BlueArhieveAPI.fetch_ba_character_list(reload_database)
                self.fetch_students(True)

            if generate_excel:
                rarity = ['3⭐', '4⭐', '5⭐', 'UE30', 'UE40', 'UE50']
                rarity.reverse()

                levels = [str(i + 1) for i in range(9)]
                levels_ex = levels[:4]
                levels.append('M')
                levels.reverse()
                levels_ex.append('M')
                levels_ex.reverse()

                filename = r'resources/public/excel/students.xlsx' # r'resources/public/excel/students-vers-' + str(datetime.date.today()) + '.xlsx'
                students = [[name, rarity, levels_ex, levels, levels, levels] for name in sr.ba_character_names]
                gen_excel(['Student', 'Rarity', 'EX Skill'
                           , 'Normal Skill', 'Passive Skill', 'Sub Skill', 'Bond'], students, filename)
                
                # try:
                #     xlsx_to_pdf(r'resources/public/excel/students.xlsx', r'resources/public/excel/students.pdf')

                #     pdf_bytes = open(r'resources/public/excel/students.pdf', 'rb').read()
                #     pdf_image = img2pdf.convert(pdf_bytes)

                #     with open(r'resources/public/excel/students.jpg', 'wb') as f:
                #         f.write(pdf_image)
                # except Exception as ex:
                #     self.logger.warning(f"Failed to save thumbnail, {ex}")

            await msg.edit(embed=discord.Embed(color=discord.Color.from_str('#02D3FB'), description="reload configuration completed"))
        finally:
            self.logger.debug("===== reload-configuration - complete =====")

    # endregion

    # region Blue Archive student info
    student = app_commands.Group(name="student", description="Blue Archive student info")

    @student.command(name='show', description='Show the student details')
    @app_commands.autocomplete(name=student_options)
    @app_commands.autocomplete(school=school_options)
    @app_commands.autocomplete(rarity=rarity)
    @app_commands.describe(name="The name of the student (start input name for suggestion)")
    @app_commands.describe(school="The school of students enrolled in")
    @app_commands.describe(rarity="The rarity of the student to be listed out")
    async def show_student(self, ctx: discord.Interaction|commands.Context, name: str = '*', school: str = '*', rarity: str = '*'):
        """Show the student details"""
        self.logger.debug("===== show-student - start =====")
        self.logger.debug(f"name => {name} :: school => {school} :: rarity => {rarity}")

        try:
            ctx = await context(ctx)

            embed = discord.Embed(color=discord.Color.from_str('#02D3FB'), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot | Blue Archive wiki @ bluearchive.wiki | EN")

            if not name == '*' and name not in sr.ba_character_names:
                best_match = check_similar(name, sr.ba_character_names, 80)
                if not best_match:
                    embed.description = f"Student **{name}** not found"
                    return await response(ctx, embed=embed)
                name = best_match
            
            if not school == '*' and school not in sr.ba_school_names:
                best_match = check_similar(school, sr.ba_school_names, 80)
                if not best_match:
                    embed.description = f"School **{school}** not found"
                    return await response(ctx, embed=embed)
                school = best_match
                
            if not rarity == '*' and rarity not in sr.ba_student_rarity[4:]:
                best_match = check_similar(rarity, sr.ba_student_rarity[4:], 100)
                if not best_match:
                    embed.description = f"Invalid rarity"
                    return await response(ctx, embed=embed)
                rarity = best_match
            
            school_detail = select("BA_SCHOOL", where={"NAME":"="}, conditions=(school,))[2][0] if not school == '*' else {}
            embed.color = discord.Color.from_str(school_detail.get("COLOR", '#02D3FB'))
            embed.set_thumbnail(url=school_detail.get("ICON", "https://static.miraheze.org/bluearchivewiki/d/d2/Ba_logo_large.png"))

            embeds = []
            result = None
            row_count = 0
            page_limit = 40

            if name == '*' and not school == '*' and not rarity == '*':
                # if not stated the student name but school and rarity (state all the student under given school and rarity)
                _, row_count, result = raw(f"SELECT * FROM BA_CHARACTER WHERE '{school}' LIKE '%' || C_SCHOOL || '%' and C_RARITY={rarity[0]} ORDER BY C_NAME")

                if row_count > 0:
                    embed = copy.deepcopy(embed)
                    embed.description = f"The list of {'<:ba_star_rarity:1138145813562916947>' * int(rarity[0])} students enrolled with **{school}**!"
                    embed.set_thumbnail(url=school_detail.get("ICON", None))
                    embed.add_field(name="", value="\b", inline=False)

                    std = ""
                    page_num = math.ceil(row_count / page_limit)
                    self.logger.debug(f"create {page_num} page(s) for {row_count} of record(s), at single page content limit of {page_limit}")
                    for p in range(page_num):
                        page_embed = copy.deepcopy(embed)
                        single_page_records = page_limit if len(result) >= page_limit else len(result)
                        self.logger.debug(f"Page {p + 1} with single page content limit of {single_page_records}")
                        n = math.ceil(single_page_records / 2)
                        for i in range(single_page_records):
                            row = result.pop(0)
                            std += f"`{row.get('C_NAME', None)}`\n"
                            if i + 1 in (n, single_page_records):
                                page_embed.add_field(name="", value=std)
                                std = ''
                        embeds.append(page_embed)
                else:
                    embed.set_thumbnail(url="https://media.discordapp.net/attachments/1138291374853197895/1138291406763466863/iibn0o5undha1-removebg-preview.png")
                    embed.add_field(inline=False,name="", value=f"> No record found! There might not have students under `{school}` with {'<:ba_star_rarity:1138145813562916947>' * int(rarity[0])} rarity")
                    embeds.append(embed)
            elif name == '*' and school == '*' and not rarity == '*':
                # if not stated the student name and school but rarity (state all the students with given rarity)
                _, row_count, result = raw(f"SELECT C_NAME, S.NAME AS C_SCHOOL FROM BA_CHARACTER C INNER JOIN BA_SCHOOL S ON S.NAME LIKE '%' || C.C_SCHOOL || '%' WHERE C_RARITY = '{rarity[0]}' ORDER BY C_SCHOOL, C_NAME")

                def get_school(item):
                    return item["C_SCHOOL"]
                
                grouped_data = {school: [record["C_NAME"] for record in group] for school, group in groupby(result, get_school)}

                embed = copy.deepcopy(embed)
                embed.description=f"The list of {'<:ba_star_rarity:1138145813562916947>' * int(rarity[0])} students!"
                embed.add_field(name="", value="\b", inline=False)

                for school, students in grouped_data.items():
                    page_embed = copy.deepcopy(embed)
                    std = ""
                    n = math.ceil(len(students) / 2) if len(students) >= 10 else len(students)
                    first = True
                    for i, student in enumerate(students):
                        std += f"`{student}`\n"
                        if i+1 in (n, len(students)):
                            page_embed.add_field(name=school if first else '\b', value=std)
                            std = ""
                            first = False
                    embeds.append(page_embed)
            elif name == '*' and not school == '*' and rarity == '*':
                # if not stated the student name and rarity but school (show all the students under that school)
                _, row_count, result = raw(f"SELECT * FROM BA_CHARACTER WHERE '{school}' LIKE '%' || C_SCHOOL || '%' ORDER BY C_NAME")

                embed = copy.deepcopy(embed)
                embed.description=f"The list of students enrolled with **{school}**~"
                embed.add_field(name="", value="\b", inline=False)

                if row_count > 0:
                    std = ""
                    n = math.ceil(row_count / 2)
                    for i in range(row_count):
                        std += f"`{result.pop(0).get('C_NAME', None)}`\n"
                        if i + 1 in (n, row_count):
                            embed.add_field(name="", value=std)
                            std = ''

                    embeds.append(embed)
                else:
                    embed.set_thumbnail(url="https://media.discordapp.net/attachments/1138291374853197895/1138291406763466863/iibn0o5undha1-removebg-preview.png")
                    embed.add_field(inline=False,name="", value=f"> No record found! There might not have students under `{school}`")
                    embeds.append(embed)
            elif not name == '*':
                # assume name only
                _, row_count, result = select('BA_CHARACTER', where={"C_NAME":"="}, conditions=(name))
                profile = result[0]
                embed = copy.deepcopy(embed)
                embed.title = f"[{'<:ba_star_rarity:1138145813562916947>' * int(profile.get('C_RARITY'))}] {profile.get('C_NAME')}"

                ## Row 1 
                embed.add_field(name='Details', value=f"**Affiliation:** {profile.get('C_SCHOOL')}\n**Type:** {profile.get('C_COMBAT_CLASS')}\n**Role/Position: **{profile.get('C_ROLE')}/{profile.get('C_POSITION')}")
                embed.add_field(name='_ _', value=f"<:ba_attack_type:1138147454240432128> **Damage Type:** {profile.get('C_ATTACK_TYPE')}\n<:ba_cover:1138147504769208362> **Armor Type:** {profile.get('C_ARMOR_TYPE')}\n**Weapon: **{profile.get('C_WEAPON')}")
                embed.add_field(name='', value='\b')

                ## Row 2 Environment Preference
                embed.add_field(name='Environment Preference', value=f"<:ba_urban_welfare:1138146891893317662> **Urban:** {self.get_preference_by_rating(profile.get('C_URBAN'))} \n<:ba_outdoors_welfare:1138146969118838865> **Outdoors:** {self.get_preference_by_rating(profile.get('C_OUTDOORS'))} \n<:ba_indoors_welfare:1138147045601980597> **Indoors:** {self.get_preference_by_rating(profile.get('C_INDOORS'))} ")
                embed.add_field(name='_ _', value=f"**Cover:** {profile.get('C_COVER')}")
                embed.add_field(name='', value='\b')

                ## Row 3 EX, Normal, Passive, Sub Skills
                # Fetch skills
                skills = api.BlueArhieveAPI.get_student_skill(name)
                if not len(skills) == 0:
                    content = ''
                    for skill in skills:
                        ski = skills.get(skill)
                        cost = ski.get("cost")
                        content += f'**{skill}: {ski.get("name")}{" [" + cost.replace(" ", ": ") + "]" if cost else ""}**\n{ski.get("description")}\n\n'
                    embed.add_field(name='Skills', value=content, inline=False)

                ### Row n
                embed.add_field(name='Introduced at', value=profile.get('C_RELEASE_DATE'), inline=False)

                school_detail = raw(f"SELECT COLOR FROM BA_SCHOOL WHERE NAME LIKE '%' || '{profile.get('C_SCHOOL')}' || '%'")[2][0]

                avatar = profile.get('C_AVATAR', '').replace('40px','120px')
                avatar = avatar if avatar.startswith(('http://', 'https://')) else f"https:{avatar}"
                embed.set_thumbnail(url=avatar)
                embed.color = discord.Color.from_str(school_detail.get("COLOR", '#02D3FB'))

                embeds.append(embed)
            else:
                # fetch all the characters and return into a page
                _, row_count, result = select('BA_CHARACTER')
                embed = copy.deepcopy(embed)
                embed.description=f"Here is all the students in Kivotos ~"

                std = ""
                page_num = math.ceil(row_count / page_limit)
                self.logger.debug(f"create {page_num} page(s) for {row_count} of record(s), at single page content limit of {page_limit}")
                for p in range(page_num):
                    page_embed = copy.deepcopy(embed)
                    single_page_records = page_limit if len(result) >= page_limit else len(result)
                    self.logger.debug(f"Page {p + 1} with single page content limit of {single_page_records}")
                    n = math.ceil(single_page_records / 2)
                    for i in range(single_page_records):
                        row = result.pop(0)
                        std += f"`{row.get('C_NAME', None)}, {row.get('C_SCHOOL', None)}`\n"
                        if i + 1 in (n, single_page_records):
                            page_embed.add_field(name="", value=std)
                            std = ''
                    embeds.append(page_embed)
            
            if len(embeds) == 1:
                return await response(ctx,embed=embeds[0])
            elif len(embeds) > 1:
                return await page.EmbedPaginator(timeout=180).start(ctx, pages=embeds)

            embed=copy.deepcopy(embed)
            embed.set_thumbnail(url="https://media.discordapp.net/attachments/1138291374853197895/1138291406763466863/iibn0o5undha1-removebg-preview.png")
            embed.description="Seem like I failed retrieve the data based on your request \nbut yooo no worry, it's not your fault, most probably due to failure during data fetching"
            await response(ctx,embed=embed)
        finally:
            self.logger.debug("===== show-student - complete =====")

    def get_preference_by_rating(self, rate: str) -> str:
        if rate == 'S':
            return '<:ba_mood_s:1138147122114478150>'
        elif rate == 'B':
            return '<:ba_mood_b:1138147167924670475>'
        elif rate == 'D':
            return '<:ba_mood_d:1138147209469231175>'
        return ''
    # endregion
    
    # region for my student
    mystudent = app_commands.Group(name="mystudent", description="Blue Archive student status prior to ease borrow")

    @mystudent.command(name='list', description='List all students')
    async def list_my_student(self, ctx: discord.Interaction|commands.Context):
        """List all students"""
        self.logger.debug("===== list-mystudent - start =====")

        try:
            ctx = await context(ctx)

            embed = discord.Embed(color=discord.Color.from_rgb(255, 170, 204), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot @ {ctx.guild.name} | EN")

            row, students = raw(f"SELECT C_NAME, RARITY, SKILL_EX, SKILL_NORMAL, SKILL_PASSIVE, SKILL_SUB, BOND FROM BA_MEMBER_STUDENTS WHERE MEMBER_ID = '{ctx.author.id}'")[1:]
            if row == 0:
                embed.description = f"There has no records of your students available, use **/mystudent add** to add stat of your student"
                return await response(ctx,embed=embed)
            embeds = []
            for s in students:
                self.logger.debug(s)
                std = ''
                page_limit = 30
                page_embed = copy.deepcopy(embed)
                page_embed.description = f"Use **/mystudent update** to update your students and **/mystudent remove** to remove added student."
                page_embed.add_field(name="", value="\b", inline=False)
                single_page_records = page_limit if len(students) >= page_limit else len(students)
                n = math.ceil(single_page_records / 2)
                for i in range(single_page_records):
                    row = students.pop(0)
                    std += f"⤖ `{row.get('RARITY', '').ljust(4, ' ')} {row.get('SKILL_EX', '')}{row.get('SKILL_NORMAL', '')}{row.get('SKILL_PASSIVE', '')}{row.get('SKILL_SUB', '')}{'('+str(row.get('BOND', ''))+')' if not row.get('BOND', '') == '' else ''}` **{row.get('C_NAME', '')}**\n"
                    if i + 1 in (n, single_page_records):
                        page_embed.add_field(name="", value=std)
                        std = ''
                page_embed.add_field(name="", value="\b", inline=False)
                page_embed.add_field(name="", value="Please ensure that your information is synchronized with your in-game account to avoid any issues or discrepancies !", inline=False)
                embeds.append(page_embed)
            if len(embeds) == 1:
                return await response(ctx,embed=embeds[0], ephemeral=True)
            elif len(embeds) > 1:
                return await page.EmbedPaginator(ephemeral=True, timeout=180).start(ctx, pages=embeds)
        finally:
            self.logger.debug("===== list-mystudent - complete =====") 

    @mystudent.command(name='add', description='Add student into your collection of students')
    @app_commands.autocomplete(rarity=rarity_full)
    @app_commands.autocomplete(name=student_options)
    @app_commands.autocomplete(ex_skill=skill_level_options)
    @app_commands.autocomplete(normal_skill=skill_level_options)
    @app_commands.autocomplete(passive_skill=skill_level_options)
    @app_commands.autocomplete(sub_skill=skill_level_options)
    @app_commands.describe(rarity="The rarity of student (e.g. UE50)")
    @app_commands.describe(name="The name of the student (start input name for suggestion)")
    @app_commands.describe(ex_skill="The skill level of ex skill")
    @app_commands.rename(ex_skill="ex-skill")
    @app_commands.describe(normal_skill="The skill level of normal skill")
    @app_commands.rename(normal_skill="normal-skill")
    @app_commands.describe(passive_skill="The skill level of passive skill")
    @app_commands.rename(passive_skill="passive-skill")
    @app_commands.describe(sub_skill="The skill level of sub skill")
    @app_commands.rename(sub_skill="sub-skill")
    @app_commands.describe(bond="relationship level")
    async def add_my_student(self, ctx: discord.Interaction|commands.Context, name: str, rarity: str, ex_skill: str, normal_skill: str, passive_skill: str, sub_skill: str, bond: int):
        """Add student into your collection of students"""
        self.logger.debug("===== add-mystudent - start =====")
        self.logger.debug(f"name => {name} :: rarity => {rarity} :: ex_skill => {ex_skill} :: normal_skill => {normal_skill} :: passive_skill => {passive_skill} :: sub_skill => {sub_skill} :: bond => {bond}")

        try:
            ctx = await context(ctx)

            embed = discord.Embed(color=discord.Color.from_rgb(255, 170, 204), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot @ {ctx.guild.name} | EN")

            if name == '*':
                embed.description = f'Hey!! You!! Don\'t you try to piplup me'
                return await response(ctx,embed=embed)

            ## Step 1: check if the user have BA profile, if not create one associate, ask them to register
            r, mem = raw(f"SELECT MEMBER_NICKNAME, MEMBER_IGN FROM BA_GUILD_MEMBER WHERE MEMBER_ID = {ctx.author.id} AND SERVER_ID = {ctx.guild.id}")[1:]
            if r == 0:
                embed.description = f"Hmmm, seem like you had not been registered as member of any club\nYou may look for any of the manager or server admin for assistance"
                return await response(ctx,embed=embed)

            member = mem[0].get('MEMBER_IGN', ctx.author.name)
            ## Step 2: check if the user have this character in his collection, if yes, prompt if they want update
            r, result = raw(f"SELECT RARITY, SKILL_EX, SKILL_NORMAL, SKILL_PASSIVE, SKILL_SUB, BOND FROM BA_MEMBER_STUDENTS WHERE MEMBER_ID = {ctx.author.id} AND C_NAME = '{name}'")[1:]
            if not r == 0:
                tembed = copy.deepcopy(embed)
                tembed.description = f"Hi {member} sensei,\nSeem like you had **{name}** in your student list\nWould you like to update it?\nCurrent skill level:\nEx - {result[0].get('SKILL_EX')}, Normal - {result[0].get('SKILL_NORMAL')}, Passive - {result[0].get('SKILL_PASSIVE')}, Sub - {result[0].get('SKILL_SUB')}, Bond - {result[0].get('BOND')}\n\nReact on ✅ to overwrite, react on ❌ to skip."
                msg: discord.Message = await response(ctx,embed=tembed) 
                await msg.add_reaction("✅")
                await msg.add_reaction("❌")
                def is_owner(reaction, user):
                    if reaction.message.id == msg.id and not user.id == ctx.author.id:
                        return False
                    if reaction.message.id == msg.id and user.id == ctx.author.id and str(reaction.emoji) == '❌':
                        raise CancelledAction("canceled, user reacted with ❌")
                    return reaction.message.id == msg.id and user.id == ctx.author.id and str(reaction.emoji) == '✅'
                try:
                    await self.bot.wait_for('reaction_add', check=is_owner)
                except:
                    self.logger.debug("User reacted '❌', continue for next configuration..")
                    tembed.description = f"Nothing reflected"
                    return await response(ctx,embed=tembed)
                finally:
                    await msg.delete()

            ## Step 3: insert record into database
            affect = raw(f"INSERT OR REPLACE INTO BA_MEMBER_STUDENTS(MEMBER_ID,C_NAME,RARITY,SKILL_EX,SKILL_NORMAL,SKILL_PASSIVE,SKILL_SUB,BOND) VALUES('{ctx.author.id}','{name}','{rarity}','{ex_skill}','{normal_skill}','{passive_skill}','{sub_skill}','{bond}')")[:1]
            self.logger.debug(f"Update student {name}'s skills for {member} to rarity:{rarity}, ex:{ex_skill}, normal:{normal_skill}, passive:{passive_skill}, sub:{sub_skill}, bond:{bond}")
            embed.description = f"Stat for **{name}** had been updated!!"
            embed.add_field(name='', value='\b', inline=False)
            embed.add_field(name=f"{name}'s skills", value=f"⤖ `{rarity.ljust(4, ' ')} {ex_skill}{normal_skill}{passive_skill}{sub_skill}({bond})`", inline=False)
            embed.add_field(name='', value='\b', inline=False)

            await response(ctx,embed=embed)
        finally:
            self.logger.debug("===== add-mystudent - complete =====") 

    batch = app_commands.Group(name="batch-add", description="Add student into your collection of students in a batch", parent=mystudent)

    @batch.command(name='download', description='Donwload the template excel for student batch update')
    async def batch_add_student(self, ctx: discord.Interaction|commands.Context):
        """Add student into your collection of students in a batch"""
        self.logger.debug("===== batch-add-mystudent - start =====")

        try:
            ctx = await context(ctx, ephemeral=True)

            embed = discord.Embed(color=discord.Color.from_rgb(255, 170, 204), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot @ {ctx.guild.name} | EN")

            ## Step 1: check if the user have BA profile, if not create one associate, ask them to register
            if not self.owner_only(ctx):
                r, mem = raw(f"SELECT MEMBER_NICKNAME, MEMBER_IGN FROM BA_GUILD_MEMBER WHERE MEMBER_ID = {ctx.author.id} AND SERVER_ID = {ctx.guild.id}")[1:]
                if r == 0:
                    embed.description = f"Hmmm, seem like you had not been registered as member of any club\nYou may look for any of the manager or server admin for assistance"
                    return await response(ctx,embed=embed)

            embed.description = f"Download the excel, fill in and upload with **/mystudent batch-add upload**"
            return await response(ctx, embed=embed, file=discord.File(r'resources/public/excel/students.xlsx'))
        finally:
            self.logger.debug("===== batch-add-mystudent - complete =====") 

    @batch.command(name='upload', description='Upload the filled excel for perform batch update of student info')
    @app_commands.describe(excel="The completed excel to be batch update")
    async def batch_add_student(self, ctx: discord.Interaction|commands.Context, excel: discord.Attachment = None):
        """Upload the filled excel for perform batch update of student info"""
        self.logger.debug("===== batch-add-mystudent - start =====")

        try:
            ctx = await context(ctx, ephemeral=True)

            embed = discord.Embed(color=discord.Color.from_rgb(255, 170, 204), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot @ {ctx.guild.name} | EN")

            ## Step 1: check if the user have BA profile, if not create one associate, ask them to register
            if not self.owner_only(ctx):
                r, mem = raw(f"SELECT MEMBER_NICKNAME, MEMBER_IGN FROM BA_GUILD_MEMBER WHERE MEMBER_ID = {ctx.author.id} AND SERVER_ID = {ctx.guild.id}")[1:]
                if r == 0:
                    embed.description = f"Hmmm, seem like you had not been registered as member of any club\nYou may look for any of the manager or server admin for assistance"
                    return await response(ctx,embed=embed)

            if excel == None:
                embed.description = f"Aren't you forget something behind?"
                return await response(ctx,embed=embed)

            # Load the Excel file
            file: discord.File = await excel.to_file()
            workbook = openpyxl.load_workbook(file.fp)

            worksheet = workbook.active
                
            # raw(f'DELETE FROM BA_MEMBER_STUDENTS WHERE MEMBER_ID = {ctx.author.id}')
            header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
            queries = ""
            updated = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if all(cell_value is not None for cell_value in row):
                    updated.append({k:v for k,v in zip(header, row)})
                    queries += f"INSERT OR REPLACE INTO `BA_MEMBER_STUDENTS`(MEMBER_ID,C_NAME,RARITY,SKILL_EX,SKILL_NORMAL,SKILL_PASSIVE,SKILL_SUB) VALUES ('{ctx.author.id}','{row[0]}','{row[1]}','{row[2]}','{row[3]}','{row[4]}','{row[5]}');\n"
            workbook.close()
            execute_script(queries)

            embeds = []
            for s in updated:
                std = ''
                page_limit = 30
                page_embed = copy.deepcopy(embed)
                page_embed.description = f"The students had been updated!!\nUse **/mystudent update** to update your students and **/mystudent remove** to remove added student."
                page_embed.add_field(name="", value="\b", inline=False)
                single_page_records = page_limit if len(updated) >= page_limit else len(updated)
                n = math.ceil(single_page_records / 2)
                for i in range(single_page_records):
                    row = updated.pop(0)
                    std += f"⤖ `{row.get('Rarity', '').ljust(4, ' ')} {row.get('EX Skill', '')}{row.get('Normal Skill', '')}{row.get('Passive Skill', '')}{row.get('Sub Skill', '')}` **{row.get('Student', '')}**\n"
                    if i + 1 in (n, single_page_records):
                        page_embed.add_field(name="", value=std)
                        std = ''
                page_embed.add_field(name="", value="\b", inline=False)
                page_embed.add_field(name="", value="Please ensure that your information is synchronized with your in-game account to avoid any issues or discrepancies !", inline=False)
                embeds.append(page_embed)
            if len(embeds) == 1:
                return await response(ctx,embed=embeds[0], ephemeral=True)
            elif len(embeds) > 1:
                return await page.EmbedPaginator(ephemeral=True, timeout=180).start(ctx, pages=embeds)
            else:
                embed.description = f"Nothing to update"

            await response(ctx,embed=embed)
        finally:
            self.logger.debug("===== batch-add-mystudent - complete =====") 

    @mystudent.command(name='update', description='Update student info from existing student collections')
    @app_commands.autocomplete(rarity=rarity_full)
    @app_commands.autocomplete(student=mystudent_options)
    @app_commands.autocomplete(ex_skill=skill_level_options)
    @app_commands.autocomplete(normal_skill=skill_level_options)
    @app_commands.autocomplete(passive_skill=skill_level_options)
    @app_commands.autocomplete(sub_skill=skill_level_options)
    @app_commands.describe(student="The name of the student (start input name for suggestion)")
    @app_commands.describe(rarity="The rarity of student (e.g. UE50)")
    @app_commands.describe(ex_skill="The skill level of ex skill")
    @app_commands.rename(ex_skill="ex-skill")
    @app_commands.describe(normal_skill="The skill level of normal skill")
    @app_commands.rename(normal_skill="normal-skill")
    @app_commands.describe(passive_skill="The skill level of passive skill")
    @app_commands.rename(passive_skill="passive-skill")
    @app_commands.describe(sub_skill="The skill level of sub skill")
    @app_commands.rename(sub_skill="sub-skill")
    @app_commands.describe(bond="relationship level")
    async def update_my_student(self, ctx: discord.Interaction|commands.Context, student: str, rarity: str = None, ex_skill: str = None, normal_skill: str = None, passive_skill: str = None, sub_skill: str = None, bond: int = None):
        """Update student info from existing student collections"""
        self.logger.debug("===== update-mystudent - start =====")

        try:
            ctx = await context(ctx)

            embed = discord.Embed(color=discord.Color.from_rgb(255, 170, 204), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot @ {ctx.guild.name} | EN")

            ## Step 1: check if the user have BA profile, if not create one associate, ask them to register
            r, mem = raw(f"SELECT MEMBER_NICKNAME, MEMBER_IGN FROM BA_GUILD_MEMBER WHERE MEMBER_ID = {ctx.author.id} AND SERVER_ID = {ctx.guild.id}")[1:]
            if r == 0:
                embed.description = f"Hmmm, seem like you had not been registered as member of any club\nYou may look for any of the manager or server admin for assistance"
                return await response(ctx,embed=embed)

            member = mem[0].get('MEMBER_IGN', ctx.author.name)
            
            r, result = raw(f"SELECT RARITY, SKILL_EX, SKILL_NORMAL, SKILL_PASSIVE, SKILL_SUB, BOND FROM BA_MEMBER_STUDENTS WHERE MEMBER_ID = {ctx.author.id} AND C_NAME = '{student}'")[1:]
            if not r == 0:
                changes = {
                    result[0].get('RARITY') : rarity,
                    result[0].get('SKILL_EX') : ex_skill,
                    result[0].get('SKILL_NORMAL') : normal_skill,
                    result[0].get('SKILL_PASSIVE') : passive_skill,
                    result[0].get('SKILL_SUB') : sub_skill,
                    result[0].get('BOND') : bond
                }
                changes = {key: value for key, value in changes.items() if value is not None}
                tembed = copy.deepcopy(embed)
                tembed.description = f"Hi {student} sensei,\nWould you like to update the stat for {student} with updates below?\nRarity: {result[0].get('RARITY')} ➭ {default_if_empty(rarity, result[0].get('RARITY'))}\nSkill level:\nEx - {result[0].get('SKILL_EX')} ➭ {default_if_empty(ex_skill, result[0].get('SKILL_EX'))}, Normal - {result[0].get('SKILL_NORMAL')} ➭ {default_if_empty(normal_skill, result[0].get('SKILL_NORMAL'))}, Passive - {result[0].get('SKILL_PASSIVE')} ➭ {default_if_empty(passive_skill, result[0].get('SKILL_PASSIVE'))}, Sub - {result[0].get('SKILL_SUB')} ➭ {default_if_empty(sub_skill, result[0].get('SKILL_SUB'))}, Bond - {result[0].get('BOND')} ➭ {default_if_empty(bond, result[0].get('BOND'))}\n\nReact on ✅ to overwrite, react on ❌ to skip."
                msg: discord.Message = await response(ctx,embed=tembed) 
                await msg.add_reaction("✅")
                await msg.add_reaction("❌")
                def is_owner(reaction, user):
                    if reaction.message.id == msg.id and not user.id == ctx.author.id:
                        return False
                    if reaction.message.id == msg.id and user.id == ctx.author.id and str(reaction.emoji) == '❌':
                        raise CancelledAction("canceled, user reacted with ❌")
                    return reaction.message.id == msg.id and user.id == ctx.author.id and str(reaction.emoji) == '✅'
                try:
                    await self.bot.wait_for('reaction_add', check=is_owner)
                except CancelledAction:
                    self.logger.debug("User reacted '❌', continue for next configuration..")
                    tembed.description = f"Nothing reflected"
                    return await response(ctx,embed=tembed)
                except Exception as e:
                    self.logger.error(e)
                    tembed.description = f"Nothing reflected, error occurs during execution"
                    return await response(ctx,embed=tembed)
                finally:
                    await msg.delete()

                ## Update into database
                affect = raw(f"INSERT OR REPLACE INTO BA_MEMBER_STUDENTS(MEMBER_ID,C_NAME,RARITY,SKILL_EX,SKILL_NORMAL,SKILL_PASSIVE,SKILL_SUB,BOND) VALUES('{ctx.author.id}','{student}','{default_if_empty(rarity, result[0].get('RARITY'))}','{default_if_empty(ex_skill, result[0].get('SKILL_EX'))}','{default_if_empty(normal_skill, result[0].get('SKILL_NORMAL'))}','{default_if_empty(passive_skill, result[0].get('SKILL_PASSIVE'))}','{default_if_empty(sub_skill, result[0].get('SKILL_SUB'))}','{default_if_empty(bond, result[0].get('BOND'))}')")[:1]
                self.logger.debug(f"Update student {student}'s skills for {member} to rarity:{rarity}, ex:{ex_skill}, normal:{normal_skill}, passive:{passive_skill}, sub:{sub_skill}, bond:{bond}")
                embed.description = f"Stat for **{student}** had been updated!!"
                std = '\n'.join([str(k) + ' ➭ ' +  str(v) for k, v in changes.items()])
                embed.add_field(name="", value=std, inline=False)
            else:
                embed.description = f"Student not found"
            await response(ctx,embed=embed)
        finally:
            self.logger.debug("===== update-mystudent - complete =====") 

    @mystudent.command(name='remove', description='Remove student from your student collections')
    @app_commands.autocomplete(student=mystudent_options)
    @app_commands.describe(student="The name of the student (start input name for suggestion)")
    async def remove_my_student(self, ctx: discord.Interaction|commands.Context, student: str):
        """Remove student from your student collections"""
        self.logger.debug("===== remove-mystudent - start =====")

        try:
            ctx = await context(ctx)

            embed = discord.Embed(color=discord.Color.from_rgb(255, 170, 204), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot @ {ctx.guild.name} | EN")

            ## Step 1: check if the user have BA profile, if not create one associate, ask them to register
            r, mem = raw(f"SELECT MEMBER_NICKNAME, MEMBER_IGN FROM BA_GUILD_MEMBER WHERE MEMBER_ID = {ctx.author.id} AND SERVER_ID = {ctx.guild.id}")[1:]
            if r == 0:
                embed.description = f"Hmmm, seem like you had not been registered as member of any club\nYou may look for any of the manager or server admin for assistance"
                return await response(ctx,embed=embed)

            member = mem[0].get('MEMBER_IGN', ctx.author.name)

            r, result = raw(f"SELECT RARITY, SKILL_EX, SKILL_NORMAL, SKILL_PASSIVE, SKILL_SUB FROM BA_MEMBER_STUDENTS WHERE MEMBER_ID = {ctx.author.id} AND C_NAME = '{student}'")[1:]
            if not r == 0:
                tembed = copy.deepcopy(embed)
                tembed.description = f"Hi {student} sensei,\nAre you sure you want to remove {student}? ⚠️ Please note that this action is irreversible.\n\nReact on ✅ to remove, react on ❌ to cancel."
                msg: discord.Message = await response(ctx,embed=tembed) 
                await msg.add_reaction("✅")
                await msg.add_reaction("❌")
                def is_owner(reaction, user):
                    if reaction.message.id == msg.id and not user.id == ctx.author.id:
                        return False
                    if reaction.message.id == msg.id and user.id == ctx.author.id and str(reaction.emoji) == '❌':
                        raise CancelledAction("canceled, user reacted with ❌")
                    return reaction.message.id == msg.id and user.id == ctx.author.id and str(reaction.emoji) == '✅'
                try:
                    await self.bot.wait_for('reaction_add', check=is_owner)
                except CancelledAction:
                    self.logger.debug("User reacted '❌', continue for next configuration..")
                    tembed.description = f"Nothing reflected"
                    return await response(ctx,embed=tembed)
                except Exception as e:
                    self.logger.error(e)
                    tembed.description = f"Nothing reflected, error occurs during execution"
                    return await response(ctx,embed=tembed)
                finally:
                    await msg.delete()

                ## Update into database
                affect = raw(f"DELETE FROM BA_MEMBER_STUDENTS WHERE MEMBER_ID = '{ctx.author.id}' AND C_NAME = '{student}'")[:1]
                self.logger.debug(f"Remove {student} from {member}'s student list")
                embed.description = f"**{student}** had been removed from your list\nYou may use **/mystudent list** to check all your registered students."
            else:
                embed.description = f"Student not found"
            await response(ctx,embed=embed)
        finally:
            self.logger.debug("===== remove-mystudent - complete =====") 
    # endregion

    # region for club management including whitelist who can manage user profile
    club = app_commands.Group(name="club", description="Club management for club manager")

    manage = app_commands.Group(name="manage", description="Manage utiliy", parent=club)

    @club.command(name='create', description='Manage setting of a club')
    @app_commands.describe(id="The club id")
    @app_commands.describe(name="The club display name")
    @app_commands.describe(description="The club description")
    @app_commands.describe(member_role="The default role to be assigned to the member after member been registered")
    @app_commands.rename(member_role="member-role")
    async def create(self, ctx: discord.Interaction|commands.Context, id: str, name: str, description: str = None, member_role: discord.Role = None):
        self.logger.debug("===== create club - start =====")

        try:
            if isinstance(ctx, discord.Interaction):
                ctx = await commands.Context.from_interaction(ctx)
            if not await self.permitted_to_execute(ctx):
                return await response(ctx,embed=discord.Embed(title="", description=f"you do not have permission to execute this command", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
            
            ctx = await context(ctx, ephemeral=True)

            self.logger.debug("checking guild existance..")
            exist, result = select('BA_GUILD', ('1', 'CONFIGURED_BY'), where={"SERVER_ID":"=", "GUILD_ID":"="}, conditions=(ctx.guild.id, id), operator=('AND', ))[1:]
            if not exist == 0:
                return await response(ctx,embed=discord.Embed(color=discord.Color.from_str('#02D3FB'), description=f"Club {name} ({id}) existed, configured by {ctx.guild.get_member(result[0].get('CONFIGURED_BY')).mention}"))
            
            col = ('SERVER_ID','GUILD_ID','GUILD_NAME','CONFIGURED_BY')
            val = (ctx.guild.id, id, name, ctx.author.id)

            if description:
                col += ('GUILD_DESCRIPTION',)
                val += (description,)
            if member_role:
                col += ('DEFAULT_ROLE',)
                val += (member_role.id,)
            
            self.logger.debug("==== Insert record ====")
            query = f"INSERT INTO BA_GUILD{str(col)} VALUES{str(val)}"
            affected = raw(query)[0]
            self.logger.debug(f"{query} :: {affected} :: {True if affected == 1 else False}")
            self.logger.debug("==== Insert record ====")

            await response(ctx,embed=discord.Embed(color=discord.Color.from_str('#02D3FB'), description=f"Club {name} ({id}) had successfully configured"))
        finally:
            self.logger.debug("===== create club - complete =====") 

    # @manage.command(name='setting', description='Manage setting of a club')
    # async def setting(self, ctx: discord.Interaction|commands.Context):
    #     self.logger.debug("===== manage club setting - start =====")

    #     try:
    #         if isinstance(ctx, discord.Interaction):
    #             ctx = await commands.Context.from_interaction(ctx)
    #         if not await self.permitted_to_execute(ctx):
    #             return await response(ctx,embed=discord.Embed(title="", description=f"you do not have permission to execute this command", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
            
    #         ctx = await context(ctx, ephemeral=True)

    #         await response(ctx,content="manage club setting dummy")
    #     finally:
    #         self.logger.debug("===== manage club setting - complete =====") 

    # region manage member
    member = app_commands.Group(name="member", description="Manage member in the club", parent=club)

    @member.command(name='add', description='Add member to the club')
    @app_commands.describe(member="which club to add the member")
    @app_commands.describe(ign="which club to add the member")
    @app_commands.describe(club="which club the member join")
    @app_commands.describe(joined_at="the date member joined the club")
    @app_commands.describe(member_type="President/Vice/Manager/Member")
    @app_commands.rename(ign="in-game-name")
    @app_commands.rename(joined_at="joined-date")
    @app_commands.rename(member_type="member-type")
    @app_commands.autocomplete(club=clubs_options)
    @app_commands.autocomplete(member_type=member_type_options)
    async def add_member(self, ctx: discord.Interaction|commands.Context, member: discord.Member, ign: str, club: int, joined_at: str = str(datetime.date.today()), member_type: str = 'Member'):
        self.logger.debug("===== add member - start =====")

        try:

            if isinstance(ctx, discord.Interaction):
                ctx = await commands.Context.from_interaction(ctx)
            if not await self.permitted_to_execute(ctx):
                return await response(ctx,embed=discord.Embed(title="", description=f"you do not have permission to execute this command", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
            
            ctx = await context(ctx, ephemeral=True)

            self.logger.debug("==== Insert record ====")
            query = f"INSERT OR REPLACE INTO BA_GUILD_MEMBER(SERVER_ID,GUILD_ID,MEMBER_ID,MEMBER_IGN,MEMBER_JOINED_DATE,MEMBER_TYPE) VALUES('{ctx.guild.id}','{club}','{member.id}','{ign}','{get_date(joined_at)}',UPPER('{member_type}'));"
            affected = raw(query)[0]
            self.logger.debug(f"{query} :: {affected} :: {True if affected == 1 else False}")
            self.logger.debug("==== Insert record ====")

            guildname = raw(f"SELECT GUILD_NAME, DEFAULT_ROLE FROM BA_GUILD WHERE GUILD_ID = {club} AND SERVER_ID = {ctx.guild.id}")[2][0]

            # Assign the role here if any
            assign_role = guildname.get('DEFAULT_ROLE')
            if assign_role:
                role = ctx.guild.get_role(int(assign_role))
                self.logger.debug(f"Assign role to user {role.name}")
                await member.add_roles(role)

            await response(ctx,embed=discord.Embed(color=discord.Color.from_str('#02D3FB'), description=f"**{ign}** ({member.mention}) chillin at **{guildname.get('GUILD_NAME')}**"))
            
        finally:
            self.logger.debug("===== add member - complete =====") 

    @member.command(name='batch-add', description='(batch) Add member to the club')
    @app_commands.describe(excel="the member list to be processed")
    async def add_member_batch(self, ctx: discord.Interaction|commands.Context, excel: discord.Attachment):
        self.logger.debug("===== batch add member - start =====")

        try:

            if isinstance(ctx, discord.Interaction):
                ctx = await commands.Context.from_interaction(ctx)
            if not await self.permitted_to_execute(ctx):
                return await response(ctx,embed=discord.Embed(title="", description=f"you do not have permission to execute this command", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
            
            ctx = await context(ctx)

            group = {}
            dataframe1 = pd.read_excel(await excel.read())
            for index, row in dataframe1.iterrows():
                username = row['discord name'].split("#")
                user = discord.utils.get(ctx.guild.members, name=username[0], discriminator=username[1]) if len(username) > 1 else discord.utils.get(ctx.guild.members, name=username[0])
                userid = user.id if not user == None else user
                if not userid:
                    # send alert here
                    continue
                # self.logger.debug(f"{user.mention if not user == None else user} :: {row['IGN']} :: {row['club']} :: {row['club id']}")
                add_to_nested_dict(group, [(row['club id'], row['club'])], (user.id, row['IGN']), as_list=True)

            print(group)
            running_result = {'success':0, 'skipped':0, 'failed': 0}
            skipped = []
            for k, v in group.items():
                r, result = raw(f"SELECT 1 FROM BA_GUILD WHERE GUILD_ID = {k[0]} AND SERVER_ID = {ctx.guild.id}")[1:]
                if r == 0:
                    r, result = raw(f"SELECT 1 FROM BA_GUILD WHERE upper(GUILD_ID) = upper('{k[1]}') AND SERVER_ID = {ctx.guild.id}")[1:]
                    if r == 0:
                        msg: discord.Message = await response(ctx,embed=discord.Embed(title="", description=f"The club, **{k[1]}** with id **{k[0]}** not yet configured"+
                                                                                                           "\nWould you like to configure a club with this name and id?"+
                                                                                                           "\n\nReact on ✅ to auto configure, react on ❌ to skip for the current configuration.", color=discord.Color.from_rgb(255, 170, 204)), 
                                                            ) 
                        await msg.add_reaction("✅")
                        await msg.add_reaction("❌")
                        def is_owner(reaction, user):
                            if reaction.message.id == msg.id and not user.id == ctx.author.id:
                                return False
                            if reaction.message.id == msg.id and user.id == ctx.author.id and str(reaction.emoji) == '❌':
                                raise CancelledAction("canceled, user reacted with ❌")
                            return reaction.message.id == msg.id and user.id == ctx.author.id and str(reaction.emoji) == '✅'
                        try:
                            await self.bot.wait_for('reaction_add', check=is_owner)
                            ## Proceed to configure the club
                            affect = raw(f"INSERT INTO BA_GUILD(SERVER_ID,GUILD_ID,GUILD_NAME,CONFIGURED_BY) VALUES('{ctx.guild.id}','{k[0]}','{k[1]}','{ctx.author.id}')")[:1]
                            self.logger.debug(f"Create club configuration :: {affect}")
                            self.fetch_club(True)
                        except CancelledAction:
                            self.logger.debug("User reacted '❌', continue for next configuration..")
                            skipped.append(k)
                            add_to_nested_dict(running_result, ['skipped'], running_result['skipped'] + len(v))
                        except Exception as e:
                            self.logger.error(e)
                            skipped.append(k)
                            add_to_nested_dict(running_result, ['skipped'], running_result['skipped'] + len(v))
                        finally:
                            await msg.delete()
            for k in skipped:
                remove_nested_item(group, [k])
            
            for k,v in group.items():
                try:
                    patch = f'RUN_CLUB_MEMBER_{k[0]}_{k[1]}' 
                    patch_filename = f"{patch}_{datetime.datetime.today().date()}.sql"
                    query = f'--{patch} - {datetime.datetime.today()}\r\n'
                    for mem in v:
                        query += f"INSERT OR REPLACE INTO BA_GUILD_MEMBER (SERVER_ID,GUILD_ID,MEMBER_ID,MEMBER_IGN,MEMBER_JOINED_DATE) values('{ctx.guild.id}','{k[0]}','{mem[0]}','{mem[1]}','{datetime.datetime.today().date()}');\n"
                    path_patch = r'.\utils\database\script\patch' + f"\{patch_filename}"
                    with open(path_patch, 'w', encoding="utf-8") as f:
                        f.write(query)
                    execute_script(path_patch)
                    add_to_nested_dict(running_result, ['success'], running_result['success'] + len(v))
                except Exception as ex:
                    self.logger.error(ex)
                    add_to_nested_dict(running_result, ['failed'], running_result['failed'] + len(v))
            return await response(ctx,embed=discord.Embed(title="", description=f"Peroro {running_result['success']} success, failed: {running_result['failed']}, {running_result['skipped']} skipped", color=discord.Color.from_rgb(255, 170, 204)))
        finally:
            self.logger.debug("===== batch add member - complete =====") 

    @member.command(name='view', description='View member from the club')
    @app_commands.describe(club="which club to view, if not provided, default will returns all in records")
    @app_commands.autocomplete(club=clubs_options)
    async def view_member(self, ctx: discord.Interaction|commands.Context, club: int = 0):
        self.logger.debug("===== view member - start =====")

        try:
            ctx = await context(ctx, ephemeral=True)

            embed = discord.Embed(color=discord.Color.from_rgb(255, 170, 204), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot @ {ctx.guild.name} | EN")
            embed.set_thumbnail(url=ctx.guild.icon.url)
            result = raw(f"SELECT G.GUILD_ID, G.GUILD_NAME, M.MEMBER_ID, M.MEMBER_NICKNAME, M.MEMBER_IGN, M.MEMBER_JOINED_DATE FROM BA_GUILD_MEMBER M INNER JOIN BA_GUILD G ON M.GUILD_ID = G.GUILD_ID WHERE M.SERVER_ID = '{ctx.guild.id}' {'AND M.GUILD_ID = ' + str(club) if not club == 0 else ''}")[2]
            club_member_group = {}
            for r in result:
                add_to_nested_dict(club_member_group,[(r['GUILD_ID'], r['GUILD_NAME'])], r, as_list=True)
                
            embeds = []
            for club, members in club_member_group.items():
                std = ''
                page_limit = 30
                page_embed = copy.deepcopy(embed)
                page_embed.description = f"The are a total of **{len(members)}** members in **{club[1]}**"
                page_embed.add_field(name="", value="\b", inline=False)
                self.logger.debug(f"{club[0]} :: {club[1]} :: {len(members)}")
                single_page_records = page_limit if len(members) >= page_limit else len(members)
                n = math.ceil(single_page_records / 2)
                for i in range(single_page_records):
                    row = members.pop(0)
                    std += f"`{row.get('MEMBER_IGN', None).ljust(15, ' ')}`\n{ctx.guild.get_member(row.get('MEMBER_ID', 0)).mention}\n"
                    if i + 1 in (n, single_page_records):
                        page_embed.add_field(name="", value=std)
                        std = ''
                embeds.append(page_embed)
            if len(embeds) == 1:
                return await response(ctx,embed=embeds[0])
            elif len(embeds) > 1:
                return await page.EmbedPaginator(ephemeral=True, timeout=180).start(ctx, pages=embeds)
            else:
                embed.description = "There has no club had been configured in this server"
                return await response(ctx, embed=embed)
        finally:
            self.logger.debug("===== view member - complete =====") 

    @member.command(name='remove', description='Remove member from the club')
    @app_commands.describe(member="which member to remove from the club")
    @app_commands.autocomplete(member=club_member_option)
    async def remove_member(self, ctx: discord.Interaction|commands.Context, member: int):
        self.logger.debug("===== remove member - start =====")

        try:
            if isinstance(ctx, discord.Interaction):
                ctx = await commands.Context.from_interaction(ctx)
            if not await self.permitted_to_execute(ctx):
                return await response(ctx,embed=discord.Embed(title="", description=f"you do not have permission to execute this command", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)

            ctx = await context(ctx, ephemeral=True)

            return await response(ctx,embed=discord.Embed(title="", description=f"Nothing", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
        finally:
            self.logger.debug("===== remove member - complete =====") 

    @member.command(name='migrate', description='Migrate member from the club to another')
    @app_commands.describe(member="which member to remove from the club")
    @app_commands.describe(club="which club to migrate to")
    @app_commands.autocomplete(club=clubs_options)
    @app_commands.autocomplete(member=club_member_option)
    async def migrate_member(self, ctx: discord.Interaction|commands.Context, member: int, club: int):
        self.logger.debug("===== migrate member - start =====")

        try:
            if isinstance(ctx, discord.Interaction):
                ctx = await commands.Context.from_interaction(ctx)
            if not await self.permitted_to_execute(ctx):
                return await response(ctx,embed=discord.Embed(title="", description=f"you do not have permission to execute this command", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)

            ctx = await context(ctx, ephemeral=True)

            return await response(ctx,embed=discord.Embed(title="", description=f"Nothing", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
        finally:
            self.logger.debug("===== migrate member - complete =====") 
    # endregion

    # region whitelisting
    @manage.command(name='permit', description='Add a member or role to whitelist')
    @app_commands.describe(member="The member given permission for critical scoreboard function.")
    @app_commands.describe(role="The role given permission for critical scoreboard function.")
    @app_commands.describe(club="Which club the role/member permitted to manage, if blank will permit for any")
    @app_commands.autocomplete(club=clubs_options)
    @admin_or_owner()
    async def permit(self, ctx: discord.Interaction|commands.Context, member: discord.Member = None, role: discord.Role = None, club: int = 0):
        self.logger.debug("===== permit club whitelist - start =====")

        try:
            if isinstance(ctx, discord.Interaction):
                ctx = await commands.Context.from_interaction(ctx)
            if not await self.permitted_to_execute(ctx):
                return await response(ctx,embed=discord.Embed(title="", description=f"you do not have permission to execute this command", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)

            if member is None and role is None:
                return await response(ctx,embed=discord.Embed(title="", description=f"Either member or role should be provided to whitelisting", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
            
            ctx = await context(ctx, ephemeral=True)

            msg = ""
            col = ('SERVER_ID', 'GUILD_ID')
            val = (ctx.guild.id, club)
            if member is not None:
                permitted_members = get_value_from_nested_dict(sr.ba_club_manage_whitelist_members, [ctx.guild.id, club])
                if member.id in permitted_members:
                    msg += f"Member {member.mention} already existed in the whitelist\n"
                else:
                    add_to_nested_dict(sr.ba_club_manage_whitelist_members, [ctx.guild.id, club], member.id, as_list = True)
                    col += ('TYPE','ID')
                    val += ('1',member.id)
                    msg += f"Member {member.mention} had been added to whitelist..\n"
            if role is not None:
                permitted_roles = get_value_from_nested_dict(sr.ba_club_manage_whitelist_roles, [ctx.guild.id, club])
                if role.id in permitted_roles:
                    msg += f"Role {role.mention} already existed in the whitelist\n"
                else:
                    add_to_nested_dict(sr.ba_club_manage_whitelist_roles,  [ctx.guild.id, club], role.id, as_list = True)
                    col += ('TYPE','ID')
                    val += ('2',role.id)
                    msg += f"Role {role.mention} had been added to whitelist..\n"

            self.logger.debug("==== Insert record ====")
            query = f"INSERT OR REPLACE INTO BA_CLUB_MANAGE_WHITELIST{str(col)} VALUES{str(val)}"
            affected = raw(query)[0]
            self.logger.debug(f"{query} :: {affected} :: {True if affected == 1 else False}")
            self.logger.debug("==== Insert record ====")
            
            return await response(ctx,embed=discord.Embed(title="", description=f"{msg}", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
        finally:
            self.logger.debug("===== permit club whitelist - complete =====") 

    @manage.command(name='deny', description='Add a member or role to whitelist')
    @app_commands.describe(member="The member given permission for critical scoreboard function.")
    @app_commands.describe(role="The role given permission for critical scoreboard function.")
    @app_commands.describe(club="Which club the role/member permitted to manage, if blank will permit for any")
    @app_commands.autocomplete(club=clubs_options)
    @admin_or_owner()
    async def deny(self, ctx: discord.Interaction|commands.Context, member: discord.Member = None, role: discord.Role = None, club: int = 0):
        self.logger.debug("===== deny club whitelist - start =====")

        try:

            if isinstance(ctx, discord.Interaction):
                ctx = await commands.Context.from_interaction(ctx)
            if not await self.permitted_to_execute(ctx):
                return await response(ctx,embed=discord.Embed(title="", description=f"you do not have permission to execute this command", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)

            if member is None and role is None:
                return await response(ctx,embed=discord.Embed(title="", description=f"Either member or role should be provided to whitelisting", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
            
            ctx = await context(ctx, ephemeral=True)

            msg = ""
            col_val = {
                'SERVER_ID': ctx.guild.id, 
                'GUILD_ID': club
            }
            if member is not None:
                permitted_members = get_value_from_nested_dict(sr.ba_club_manage_whitelist_members, [ctx.guild.id, club])
                if not member.id in permitted_members:
                    msg += f"Member {member.mention} currently not in the whitelist\n"
                else:
                    col_val['TYPE'] = 1
                    col_val['ID'] = member.id
                    msg += f"Member {member.mention} had been removed from whitelist..\n"
            if role is not None:
                permitted_roles = get_value_from_nested_dict(sr.ba_club_manage_whitelist_roles, [ctx.guild.id, club])
                if not role.id in permitted_roles:
                    msg += f"Role {role.mention} currently not in the whitelist\n"
                else:
                    col_val['TYPE'] = 2
                    col_val['ID'] = role.id
                    msg += f"Role {role.mention} had been removed from whitelist..\n"
            
            self.logger.debug("==== Insert record ====")
            conditions = " AND ".join([f"{k}='{v}'" for k,v in col_val.items()])
            query = f"DELETE FROM BA_CLUB_MANAGE_WHITELIST WHERE {conditions}"
            affected = raw(query)[0]
            self.logger.debug(f"{query} :: {affected} :: {True if affected == 1 else False}")
            self.logger.debug("==== Insert record ====")

            # Refresh the whitelist
            self.fetch_whitelist(force=True)
            
            return await response(ctx,embed=discord.Embed(title="", description=f"{msg}", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)

        finally:
            self.logger.debug("===== deny club whitelist - complete =====") 

    @manage.command(name='whitelist', description='List a member or role in whitelist')
    @app_commands.describe(option="The option to list the member and/or roles whitelisted")
    @app_commands.autocomplete(option=whitelist_group_options)
    @admin_or_owner()
    async def whitelist(self, ctx: discord.Interaction|commands.Context, option: str = None):
        self.logger.debug("===== list club whitelist - start =====")

        try:

            if isinstance(ctx, discord.Interaction):
                ctx = await commands.Context.from_interaction(ctx)
            if not await self.permitted_to_execute(ctx):
                return await response(ctx,embed=discord.Embed(title="", description=f"you do not have permission to execute this command", color=discord.Color.from_rgb(255, 170, 204)), ephemeral=True)
            
            ctx = await context(ctx, ephemeral=True)

            permitted_members = get_value_from_nested_dict(sr.ba_club_manage_whitelist_members, [ctx.guild.id])
            permitted_roles = get_value_from_nested_dict(sr.ba_club_manage_whitelist_roles, [ctx.guild.id])
            embed = discord.Embed(title="", color=discord.Color.from_rgb(255, 170, 204))
            if option == 'member':
                embed.description = "Member in the whitelist are ~"
                for club, mem in permitted_members.items():
                    embed.add_field(name=club if not club == 0 else "All club", value="\n".join(ctx.guild.get_member(m).mention for m in mem) if len(mem) > 0 else "no record")
            elif option == 'role':
                embed.description = "Roles in the whitelist are ~"
                for club, role in permitted_roles.items():
                    embed.add_field(name=club if not club == 0 else "All club", value="\n".join(ctx.guild.get_role(r).mention for r in role) if len(role) > 0 else "no record")
            else:
                embed.description = "Roles and Members in the whitelist are ~"
                for club, mem in permitted_members.items():
                    embed.add_field(name=club if not club == 0 else "All club", value="\n".join(ctx.guild.get_member(m).mention for m in mem) if len(mem) > 0 else "no record")
                for club, role in permitted_roles.items():
                    embed.add_field(name=club if not club == 0 else "All club", value="\n".join(ctx.guild.get_role(r).mention for r in role) if len(role) > 0 else "no record")
            
            await response(ctx,embed=embed, ephemeral=True)
        finally:
            self.logger.debug("===== list club whitelist - complete =====") 
    # endregion

    # endregion

    # region Borrow list
    @mystudent.command(name='lookup', description='Lookup for a students among the club.')
    @app_commands.autocomplete(student=student_options)
    @app_commands.autocomplete(club=clubs_options)
    @app_commands.describe(student="which student to be lookup")
    @app_commands.describe(club="the club of members to lookup for borrow list")
    async def student_lookup(self, ctx: discord.Interaction|commands.Context, student: str, club: int = 0):
        """Lookup for a students among the club."""
        self.logger.debug("===== lookup borrow list - start =====")

        try:
            ctx = await context(ctx)

            embed = discord.Embed(color=discord.Color.from_rgb(255, 170, 204), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot @ {ctx.guild.name} | EN")

            if not club == 0:
                guild = raw(f"SELECT GUILD_NAME FROM BA_GUILD WHERE GUILD_ID = '{club}' AND SERVER_ID = {ctx.guild.id}")[2][0]
                embed.description = f"These are the members with **{student}** from **{guild.get('GUILD_NAME')}**"
            else:
                embed.description = f"These are the members with **{student}**"

            embeds = []
            if student == '*':
                embed.description = f"Listing all the available students{' in **' + guild.get('GUILD_NAME') + '**' if not club == 0 else '..'}"
                r, result = raw(f"SELECT M.C_NAME, M.MEMBER_ID, M.RARITY, M.SKILL_EX, M.SKILL_NORMAL, M.SKILL_PASSIVE, M.SKILL_SUB, M.BOND FROM BA_MEMBER_STUDENTS M LEFT JOIN BA_GUILD_MEMBER G ON M.MEMBER_ID = G.MEMBER_ID WHERE G.SERVER_ID = {ctx.guild.id} {'AND GUILD_ID = ' + str(club) if not club == 0 else ''} ORDER BY C_NAME, CASE RARITY WHEN 'UE50' THEN 1 WHEN 'UE40' THEN 2 WHEN 'UE30' THEN 3 WHEN '4⭐' THEN 4 WHEN '3⭐' THEN 5 WHEN '2⭐' THEN 6 WHEN '1⭐' THEN 7 ELSE 8 END, BOND DESC, SKILL_EX DESC, SKILL_NORMAL DESC, SKILL_PASSIVE DESC, SKILL_SUB DESC")[1:]
                group = {}
                if not r == 0:
                    for s in result:
                        add_to_nested_dict(group, [s.get('C_NAME')], s, as_list=True)
                    current_total = 0
                    page_limit = 30
                    
                    size = len(result)
                    page_embed = copy.deepcopy(embed)
                    page_embed.add_field(name="", value="\b", inline=False)
                    for index, (student_name, members) in enumerate(group.items()):
                        std = ''
                        if current_total + len(members) > page_limit or (not index == 0 and index % 2 == 0) or (index + 1 == size):
                            page_embed.add_field(name="", value="\b", inline=False)
                            embeds.append(page_embed)
                            if not (index + 1 == size):
                                page_embed = copy.deepcopy(embed)
                                page_embed.add_field(name="", value="\b", inline=False)
                                current_total = 0

                        current_total += len(members)
                        for member in members:
                            mem = ctx.guild.get_member(member.get('MEMBER_ID', 0))
                            if not mem is None:
                                std += f"⤖ `{member.get('RARITY', '').ljust(4, ' ')} {member.get('SKILL_EX', '')}{member.get('SKILL_NORMAL', '')}{member.get('SKILL_PASSIVE', '')}{member.get('SKILL_SUB', '')}{'('+str(member.get('BOND', ''))+')' if not member.get('BOND', '') == '' else ''}` **{mem.mention}**\n"
                        page_embed.add_field(name=student_name, value=std)
            else:
                r, result = raw(f"SELECT M.MEMBER_ID, M.RARITY, M.SKILL_EX, M.SKILL_NORMAL, M.SKILL_PASSIVE, M.SKILL_SUB, M.BOND FROM BA_MEMBER_STUDENTS M LEFT JOIN BA_GUILD_MEMBER G ON M.MEMBER_ID = G.MEMBER_ID WHERE M.C_NAME LIKE '{student}' AND G.SERVER_ID = {ctx.guild.id} {'AND GUILD_ID = ' + str(club) if not club == 0 else ''} ORDER BY C_NAME, CASE RARITY WHEN 'UE50' THEN 1 WHEN 'UE40' THEN 2 WHEN 'UE30' THEN 3 WHEN '4⭐' THEN 4 WHEN '3⭐' THEN 5 WHEN '2⭐' THEN 6 WHEN '1⭐' THEN 7 ELSE 8 END, BOND DESC, SKILL_EX DESC, SKILL_NORMAL DESC, SKILL_PASSIVE DESC, SKILL_SUB DESC")[1:]
                if not r == 0:
                    for s in result:
                        self.logger.debug(s)
                        std = ''
                        page_limit = 30
                        page_embed = copy.deepcopy(embed)
                        page_embed.add_field(name="", value="\b", inline=False)
                        single_page_records = page_limit if len(result) >= page_limit else len(result)
                        # n = math.ceil(single_page_records / 2)
                        n = single_page_records
                        for i in range(single_page_records):
                            row = result.pop(0)
                            mem = ctx.guild.get_member(row.get('MEMBER_ID', 0))
                            if not mem is None:
                                std += f"⤖ `{row.get('RARITY', '').ljust(4, ' ')} {row.get('SKILL_EX', '')}{row.get('SKILL_NORMAL', '')}{row.get('SKILL_PASSIVE', '')}{row.get('SKILL_SUB', '')}{'('+str(row.get('BOND', ''))+')' if not row.get('BOND', '') == '' else ''}` **{mem.mention}**\n"
                            if i + 1 in (n, single_page_records):
                                page_embed.add_field(name="", value=std)
                                std = ''
                        page_embed.add_field(name="", value="\b", inline=False)
                        page_embed.add_field(name="", value="Please ensure that your information is synchronized with your in-game account to avoid any issues or discrepancies !", inline=False)
                        embeds.append(page_embed)
            if len(embeds) == 1:
                return await response(ctx,embed=embeds[0])
            elif len(embeds) > 1:
                return await page.EmbedPaginator(ephemeral=True, timeout=180).start(ctx, pages=embeds)
            else:
                embed.add_field(name="", value="Student not found", inline=False)
                return await response(ctx,embed=embed)
        finally:
            self.logger.debug("===== lookup borrow list - complete =====") 

    @mystudent.command(name='check', description='Listdown the students of a member')
    @app_commands.describe(member="the member to listdown all the students")
    async def list_members_student(self, ctx: discord.Interaction|commands.Context, member: discord.Member):
        self.logger.debug("===== list members student - start =====")

        try:
            ctx = await context(ctx)

            embed = discord.Embed(color=discord.Color.from_rgb(255, 170, 204), timestamp=datetime.datetime.now())
            embed.set_author(name=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei！", icon_url=ctx.author.avatar)
            embed.set_footer(icon_url=self.bot.user.avatar, text=f"AiriBot @ {ctx.guild.name} | EN")

            r, mem = raw(f"SELECT MEMBER_NICKNAME, MEMBER_IGN FROM BA_GUILD_MEMBER WHERE MEMBER_ID = {member.id} AND SERVER_ID = {ctx.guild.id}")[1:]
            if r == 0:
                embed.description = f"There has no records of this member, it could be due to the member is not enrolled with any club"
                return await response(ctx,embed=embed)
            
            member_name = mem[0].get('MEMBER_IGN', member.id)

            row, students = raw(f"SELECT C_NAME, RARITY, SKILL_EX, SKILL_NORMAL, SKILL_PASSIVE, SKILL_SUB, BOND FROM BA_MEMBER_STUDENTS WHERE MEMBER_ID = '{member.id}' ORDER BY C_NAME, CASE RARITY WHEN 'UE50' THEN 1 WHEN 'UE40' THEN 2 WHEN 'UE30' THEN 3 WHEN '4⭐' THEN 4 WHEN '3⭐' THEN 5 WHEN '2⭐' THEN 6 WHEN '1⭐' THEN 7 ELSE 8 END, BOND DESC, SKILL_EX DESC, SKILL_NORMAL DESC, SKILL_PASSIVE DESC, SKILL_SUB DESC")[1:]
            if row == 0:
                embed.description = f"Seem like **{member_name}** quite lazy to update the borrow list"
                return await response(ctx,embed=embed)
            embeds = []
            for s in students:
                self.logger.debug(s)
                std = ''
                page_limit = 30
                page_embed = copy.deepcopy(embed)
                page_embed.description = f"These are **{member_name}'s** student status"
                page_embed.add_field(name="", value="\b", inline=False)
                single_page_records = page_limit if len(students) >= page_limit else len(students)
                # n = math.ceil(single_page_records / 2)
                n = single_page_records
                for i in range(single_page_records):
                    row = students.pop(0)
                    std += f"⤖ `{row.get('RARITY', '').ljust(4, ' ')} {row.get('SKILL_EX', '')}{row.get('SKILL_NORMAL', '')}{row.get('SKILL_PASSIVE', '')}{row.get('SKILL_SUB', '')}` **{row.get('C_NAME', '')}{'('+row.get('BOND', '')+')' if not row.get('BOND', '') == '' else ''}**\n"
                    if i + 1 in (n, single_page_records):
                        page_embed.add_field(name="", value=std)
                        std = ''
                page_embed.add_field(name="", value="\b", inline=False)
                # page_embed.add_field(name="", value="Please ensure that your information is synchronized with your in-game account to avoid any issues or discrepancies !", inline=False)
                embeds.append(page_embed)
            if len(embeds) == 1:
                return await response(ctx,embed=embeds[0])
            elif len(embeds) > 1:
                return await page.EmbedPaginator(ephemeral=True, timeout=180).start(ctx, pages=embeds)
        finally:
            self.logger.debug("===== list members student - complete =====") 

    # endregion

    # region profile
    @app_commands.command(name='profile', description='Display and update the profile of a member')
    @app_commands.describe(member="the member to spy the profile")
    async def profile(self, ctx: discord.Interaction|commands.Context, member: discord.Member = None):
        """Display and update the profile of a member"""
        self.logger.debug("===== profile =====")
        try:  
            ctx = await context(ctx)
            
            bot: Bot = self.bot
            user: discord.Member = ctx.author if member is None else member
            embed = discord.Embed(title=f"Hi {get_ign_or_default(ctx.author.id, ctx.author.display_name)} sensei!", 
                                description=f"", 
                                color=user.color, 
                                timestamp = datetime.datetime.now())

            row, profile = raw(f"SELECT MEMBER_IGN, MEMBER_TYPE, MEMBER_JOINED_DATE, MEMBER_FRIEND_CODE FROM BA_GUILD_MEMBER WHERE MEMBER_ID = '{user.id}';")[1:]
            if row == 0:
                embed.description = f"There has no records of this member, it could be due to the member is not enrolled with any club"
                return await response(ctx,embed=embed)
            
            # region row 1
            embed.add_field(name="Name", value=profile[0].get('MEMBER_IGN', user.display_name) if row > 0 else user.display_name)
            embed.add_field(name="Position", value=profile[0].get('MEMBER_TYPE', '-').title() if row > 0 else '-')
            embed.add_field(name="Join Date", value=profile[0].get('MEMBER_JOINED_DATE', user.joined_at.date().strftime("%d-%b-%Y")) if row > 0 else user.joined_at.date().strftime("%d-%b-%Y"))
            # endregion

            # region row 2
            embed.add_field(name="Friend Code", value=profile[0].get('MEMBER_FRIEND_CODE', "-") if row > 0 else "-")
            embed.add_field(name="", value="\b")
            embed.add_field(name="", value="\b")
            # endregion

            # region row 3
            row, average_raid = raw(f"SELECT AVG(RE_SCORE) AS AVG_SCORE, AVG(RE_RANK) AS AVG_RANK FROM BA_RAID_MEMBER RM INNER JOIN BA_RAID_RECORD BRR ON RM.RE_ID = BRR.ID WHERE RM.RE_DIFFICULTY >= '5' AND MEMBER_ID = '{user.id}';")[1:]
            avg_rank = average_raid[0].get('AVG_RANK')
            if avg_rank is not None:
                embed.add_field(name="Avg. Rank (Ins.)", value=f"{get_rank_icon(int(avg_rank))} {int(avg_rank)}")
            else:
                embed.add_field(name="Avg. Rank (Ins.)", value="-")

            row, latest_raid = raw(f"SELECT (SELECT DESCRIPTION FROM MISC WHERE TYPE = 'R_TYPE' AND NAME = UPPER(BRR.R_TYPE)) AS TYPE, BR.R_NAME || ' S-' || BRR.R_SEASON AS LAST_RAID, RE_SCORE AS LAST_RAID_SCORE, RE_RANK AS LAST_RAID_RANK FROM BA_RAID_RECORD BRR INNER JOIN BA_RAID_MEMBER RM ON RM.RE_ID = BRR.ID INNER JOIN BA_RAID BR ON BR.ID = BRR.R_ID WHERE MEMBER_ID = '{user.id}' ORDER BY BRR.R_END_DATE DESC LIMIT 1;")[1:]
            if row > 0:
                rank = int(latest_raid[0].get('LAST_RAID_RANK', 999999))
                embed.add_field(name="Latest Raid", value=f"{get_rank_icon(rank)} {rank}\n{latest_raid[0].get('TYPE')} {latest_raid[0].get('LAST_RAID')}")
            else:
                embed.add_field(name="Latest Raid", value="-")

            row, best_raid = raw(f"SELECT (SELECT DESCRIPTION FROM MISC WHERE TYPE = 'R_TYPE' AND NAME = UPPER(BRR.R_TYPE)) AS TYPE, BR.R_NAME || ' S-' || BRR.R_SEASON AS TOP_RAID, RE_SCORE AS TOP_SCORE, RE_RANK AS TOP_RANK FROM BA_RAID_MEMBER RM INNER JOIN BA_RAID_RECORD BRR ON RM.RE_ID = BRR.ID INNER JOIN BA_RAID BR ON BR.ID = BRR.R_ID WHERE MEMBER_ID = '{user.id}' ORDER BY RE_RANK LIMIT 1;")[1:]
            if row > 0:
                embed.add_field(name="Best Rank", value=f"{get_rank_icon(rank)} {rank}\n{best_raid[0].get('TYPE')} {best_raid[0].get('TOP_RAID')}")
            else:
                embed.add_field(name="Best Rank", value="-")
            # endregion

            # region row 4
            students = raw(f"SELECT MEMBER_ID, RARITY, COUNT(RARITY) AS GR_COUNT FROM BA_MEMBER_STUDENTS WHERE MEMBER_ID = '{user.id}' AND RARITY IN ('UE30', 'UE40', 'UE50') GROUP BY MEMBER_ID, RARITY ORDER BY RARITY;")[2]

            ue_std = {}
            for rarity in students:
                ue_std[rarity.get('RARITY')] = rarity.get('GR_COUNT')
                    
            embed.add_field(name="UE30", value=ue_std.get('UE30', '-'))
            embed.add_field(name="UE40", value=ue_std.get('UE40', '-'))
            embed.add_field(name="UE50", value=ue_std.get('UE50', '-'))
            # endregion

            embed.set_thumbnail(url=user.display_avatar)
            embed.set_footer(text=f"Powered by {bot.user.name}", icon_url=bot.user.avatar)
            await UpdateProfileButtonView(timeout=180).start(ctx, user, embed=embed)
        finally:
            self.logger.debug("===== profile - complete =====") 
    # endregion

def get_ign_or_default(userid, default):
    r, result = raw(f'SELECT MEMBER_IGN FROM BA_GUILD_MEMBER WHERE MEMBER_ID = {userid}')[1:]
    if not r == 0:
        return result[0].get('MEMBER_IGN', default)
    return default

def get_rank_icon(rank):
    if rank <= 15000: 
        return "<:ba_rank_plat:1136568139702882366>"
    elif rank > 15000 and rank <= 80000:
        return "<:ba_rank_gold:1136568105452179456>"
    elif rank > 80000 and rank <= 180000:
        return "<:ba_rank_silv:1136568062041137202>"
    return "<:ba_rank_brz:1136568021113126922>"
            
class ModalEditProfile(discord.ui.Modal):
    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.add_item(discord.ui.TextInput(label="In-game Name", required=False, max_length=100))
        self.add_item(discord.ui.TextInput(label="Friend Code", required=False, max_length=100))
        self.add_item(discord.ui.TextInput(label="Join Date/Joined Since", required=False, placeholder="how many days had been joined the club"))

    async def on_submit(self, interaction: discord.Interaction):
        # update into database
        ign = self.children[0].value
        ign = None if ign == '' else ign

        friendCode = self.children[1].value
        friendCode = None if friendCode == '' else friendCode
        date = self.children[2].value
        if date is not None:
            try:
                date = get_date(str((datetime.datetime.now() - datetime.timedelta(days=int(date)))))
            except:
                date = get_date(date)

        sql = []
        if ign is not None:
            sql.append(f"MEMBER_IGN = '{ign}'")
        if friendCode is not None:
            sql.append(f"MEMBER_FRIEND_CODE = '{friendCode}'")
        if date is not None and not date == '':
            sql.append(f"MEMBER_JOINED_DATE = '{date}'")

        if len(sql) > 0:
            raw(f"UPDATE BA_GUILD_MEMBER SET {','.join(sql)}, UPDATED_ON = '{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}', UPDATED_BY = UPPER('{interaction.user.display_name[:20]}') WHERE MEMBER_ID = {interaction.user.id}")

        embed = discord.Embed(color=discord.Color.from_str('#02D3FB'), title="Your profile had been updated ~")
        embed.set_author(name=f"Hi {get_ign_or_default(interaction.user.id, interaction.user.display_name)} sensei！", icon_url=interaction.user.avatar)
        embed.set_footer(icon_url=interaction.client.user.avatar, text=f"AiriBot @ {interaction.guild.name} | EN")
        embed.add_field(name="In-game Name", value=default_if_empty(ign, '-'))
        embed.add_field(name="Friend Code", value=default_if_empty(friendCode, '-'))
        embed.add_field(name="Join Date/Joined Since", value=default_if_empty(date, '-'))
        await interaction.response.send_message(embed=embed, ephemeral=True)

class UpdateProfileButtonView(discord.ui.View):
    def __init__(self, *, ephemeral: bool = True, timeout: int = None):
        self.ephemeral = ephemeral
        super().__init__(timeout=timeout)

    async def button_callback(self, interaction:discord.Interaction, button: discord.Button = None):
        if not self.user == interaction.user:
            return await interaction.response.send_message(ephemeral=True, content="You cannot edit profile for other member")
        await interaction.response.send_modal(ModalEditProfile(title="Update Profile"))

    async def start(self, ctx: discord.Interaction|commands.Context, user: discord.Member, *args, **kargs):
        if isinstance(ctx, discord.Interaction):
            ctx = await commands.Context.from_interaction(ctx)

        button = discord.ui.Button(label="Update Profile", style=ButtonStyle.primary)
        button.callback = self.button_callback
        self.add_item(button)

        self.user = user
        self.message = await ctx.send(ephemeral=self.ephemeral, view=self, *args, **kargs)

    async def on_timeout(self):
        for child in self.children:
            child.disabled = True
        self.stop()
        await self.message.edit(view=self)

class CancelledAction(Exception):
    # handle exception for cancelled action from user
    pass

async def setup(bot):
    await bot.add_cog(BlueArchive(bot))