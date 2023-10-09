from apnggif import apnggif

import os
import requests
import string
import random
import json

from dateutil.parser import parse
import validators
import polib
# from profanity_check import predict, predict_prob

import xlsxwriter
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from fuzzywuzzy import process

from utils.constant import Constant as const
import utils.logger as logger
import utils.database.connection as conn
import utils.translator as translator
import staticrunner as sr

TEMP_PATH = "\\temp\\apng_convert"

def apngtogif(url: str):
    r = requests.get(url, allow_redirects=True)
    path_to_temp = f"{os.getcwd()}{TEMP_PATH}"
    if not os.path.isdir(path_to_temp):
        os.makedirs(path_to_temp)
    tname = generateRandomStr(32)
    fpath_input = f"{path_to_temp}\\{tname}.png"
    open(fpath_input, 'wb').write(r.content)
    apnggif(png=fpath_input)
    logger.Logger().debug(f"name of file: {tname}, path: {fpath_input}")

    if os.path.exists(fpath_input):
        os.remove(fpath_input)
    return f"{path_to_temp}/{tname}.gif", f"{tname}.gif"

async def rename_file(path: str, extension: str):
    tname = None
    if validators.url(path):
        r = requests.get(path, allow_redirects=True)
        path_to_temp = f"{os.getcwd()}{TEMP_PATH}"
        if not os.path.isdir(path_to_temp):
            os.makedirs(path_to_temp)
        tname = f"{generateRandomStr(32)}.{extension}"
        path = f"{path_to_temp}\\{tname}"

        open(path, 'wb').write(r.content)

        return path, tname

    base = os.path.splitext(path)[0]
    renamed = base + "." + extension
    os.rename(path, renamed)
    
    return renamed, tname

def generateRandomStr(len: int):
    res = ''.join(random.choices(string.ascii_uppercase +
                    string.digits, k=len))
    return str(res)

def convertSeconds(seconds):
    seconds = seconds % (24 * 3600)
    hour = seconds // 3600
    seconds %= 3600
    minutes = seconds // 60
    seconds %= 60
     
    return "%d:%02d:%02d" % (hour, minutes, seconds)

def retrieve_configuration():
    data = None
    for f in const.CONF_NAME:
        path = f"{os.getcwd()}\\{f}" or f"{os.getcwd()}\\conf\\{f}"
        logger.Logger().debug(f"scanning path :: {path}")
        if not os.path.exists(path):
            continue
        f = open(f)
        data = json.load(f)
        break
    return data

def is_date(string, fuzzy=False):
    """
    Return whether the string can be interpreted as a date.

    :param string: str, string to check for date
    :param fuzzy: bool, ignore unknown tokens in string if True
    """
    try: 
        parse(string, fuzzy=fuzzy)
        return True

    except ValueError:
        return False
    
def get_date(string) -> str:
    try: 
        return str(parse(string, yearfirst=True).date())

    except ValueError:
        return ''

def set_environment_variable(key, val):
    os.environ[key] = val

def kms_get_key(id = "169926144"):
    _, row, result = conn.select(columns=("KEY",), table="KMS", where={"ID":"="}, conditions=(id,))
    if not row == 0:
        return result[0]["KEY"]
    raise IndexError(f"Key not found by given id {id}")

def is_prod() -> bool:
    return const.ENV.upper() in ("PROD", "PRODUCTION", "LIVE")

def check_send_email() -> bool:
    return is_prod() and "email" in const.DEVP_NOTIFICATION_CHANNEL

# def offensive_word_detection(text: str):
#     return predict_prob([text])

def add_to_nested_dict(nested_dict, keys, value, as_list = False):
    current_dict = nested_dict
    for key in keys[:-1]:
        current_dict = current_dict.setdefault(key, {})
    last_key = keys[-1]
    
    if as_list:
        if last_key in current_dict and isinstance(current_dict[last_key], list):
            current_dict[last_key].append(value)
        else:
            current_dict[last_key] = value if isinstance(value, list) else [value]
    else:
        current_dict[last_key] = value

def get_value_from_nested_dict(nested_dict, keys, default = None):
    current_dict = nested_dict
    for key in keys:
        if key in current_dict:
            current_dict = current_dict[key]
        else:
            return default
    return current_dict

def remove_nested_item(nested_dict, keys):
    current_dict = nested_dict
    for key in keys[:-1]:
        if key in current_dict:
            current_dict = current_dict[key]
        else:
            return
    last_key = keys[-1]
    if last_key in current_dict:
        del current_dict[last_key]

def default_if_empty(val ,default):
    return val if not val == None else default

def get_server_locale(server_id) -> str:
    return get_value_from_nested_dict(sr.StaticRunner.server_locale, [server_id], "en_US")

def __load_server_locale(servers):
    """
    Load locale configured in database
    """
    for server in servers:
        add_to_nested_dict(sr.StaticRunner.server_locale, [server.get('SERVER_ID')], server.get('LOCALE'))

def __load_server_default_role(servers):
    for server in servers:
        enabled = server.get('EN_DEF_ROLE')
        add_to_nested_dict(sr.StaticRunner.default_role_on_join, [server.get('SERVER_ID')], {"enabled":True if enabled == 1 else False, "default_role": server.get('DEF_ROLE')})
        # add_to_nested_dict(sr.StaticRunner.defaultRole, [server.get('SERVER_ID')], server.get('DEF_ROLE'))
        # add_to_nested_dict(sr.StaticRunner.setDefaultRoleOnUserJoin, [server.get('SERVER_ID')], True if enabled == 1 else False)

def load_server_setting():
    row, result = conn.select(columns=("SERVER_ID","LOCALE","EN_DEF_ROLE","DEF_ROLE"), table="DC_SERVER_SETTING")[1:]
    __load_server_locale(result)
    __load_server_default_role(result)

def load_translator():
    compile_locale_translation()
    for locale in const.SUPPORTED_LANGUAGE:
        try:
            logger.Logger().debug(f"Loading translator for locale {locale}")
            add_to_nested_dict(sr.StaticRunner.translators, [locale], translator.Translator(locale))
        except:
            logger.Logger().warning(f"Failed to load translator for locale {locale}, fall back to default en_US")
            add_to_nested_dict(sr.StaticRunner.translators, [locale], translator.Translator("en_US"))

def _gettext(msg, locale) -> str:
    try:
        t = get_value_from_nested_dict(sr.StaticRunner.translators, [locale])
        return t.translate(msg)
    except:
        t = get_value_from_nested_dict(sr.StaticRunner.translators, ['en_US'])
        return t.translate(msg)
    
def check_similar(input, list, similarity_percentage) -> str:
    best_match, similarity = process.extractOne(input, list)
    logger.Logger().debug(f"input :: {input}, best match :: {best_match}, similarity :: {similarity}")
    return best_match if similarity >= similarity_percentage else None
    
def compile_po_to_mo(po_file_path):
    # Load the .po file
    po = polib.pofile(po_file_path)

    # Compile the .po file into a .mo file
    mo_file_path = po_file_path.replace('.po', '.mo')
    po.save_as_mofile(mo_file_path)

    return mo_file_path

def compile_locale_translation():
    for language in const.SUPPORTED_LANGUAGE:
        try:
            po_file_path = f'locales\\{language}\\LC_MESSAGES\\airibot.po'
            mo_file_path = compile_po_to_mo(po_file_path)
            logger.Logger().debug(f'{language} :: {po_file_path} compiled to {mo_file_path}')
        except Exception as e:
            logger.Logger().error(e)
            logger.Logger().warning(f'Failed to compile language translation file :: {language}')

def generate_excel(headers, data_list, path):
    workbook = xlsxwriter.Workbook(path)
    worksheet = workbook.add_worksheet()

    data = [headers] + data_list
    locked_format = workbook.add_format({'locked': True})

    # Add data
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            if isinstance(value, list):
                worksheet.data_validation(row_idx, col_idx, row_idx, col_idx, {'validate': 'list', 'source': value})
                continue
        row_data = [lst for lst in row_data if not isinstance(lst, list)]
        worksheet.write_row(row_idx, 0, row_data, locked_format)
        
    worksheet.autofit()
    workbook.close()

def xlsx_to_pdf(input, output):
    # Load Excel data using openpyxl
    workbook = openpyxl.load_workbook(input)
    sheet = workbook.active

    # Create PDF using reportlab
    pdf_file = output
    c = canvas.Canvas(pdf_file, pagesize=letter)

    for row in sheet.iter_rows(values_only=True):
        c.drawString(100, 700, '\t'.join(str(cell) for cell in row))
        c.showPage()

    c.save()
    
from PySide6 import QtCore
class FileWatcher():

    fs_watcher = None
    def __init__(self, path: list| str):
        if not isinstance(path, list):
            path = [path]

        fs_watcher = QtCore.QFileSystemWatcher(path)

        fs_watcher.fileChanged.connect(self.file_changed)
        fs_watcher.directoryChanged.connect(self.directory_changed)

    def directory_changed(path):
        print('Directory Changed!!!')

    def file_changed(path):
        print('File Changed!!!')